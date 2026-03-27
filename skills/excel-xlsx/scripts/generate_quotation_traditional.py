#!/usr/bin/env python3
"""生成 Excel 报价单 - 传统工整对称风格"""

import sys
import json
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def create_border():
    """创建细边框"""
    thin = Side(style='thin', color='000000')
    return Border(top=thin, left=thin, right=thin, bottom=thin)

def create_quotation_excel(output_path, data=None):
    """创建传统风格 Excel 报价单"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    
    # ============ 页面设置 ============
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_setup.orientation = 'landscape'  # 横向打印
    # 页边距（英寸）- 参考 Word 的设置（2.5cm ≈ 1 英寸）
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.8
    ws.page_margins.bottom = 0.8
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    
    # 列宽（优化后）
    ws.column_dimensions['A'].width = 10    # Item
    ws.column_dimensions['B'].width = 25    # Description（加宽）
    ws.column_dimensions['C'].width = 10    # Qty
    ws.column_dimensions['D'].width = 15    # 条款
    
    # ============ 数据准备 ============
    company = data or {}
    customer = company.get('customer', {})
    quotation = company.get('quotation', {})
    products = company.get('products', [])
    
    # ============ 第 1 部分：标题 ============
    # 合并 A1:D1 - 公司名
    ws.merge_cells('A1:D1')
    ws['A1'] = company.get('company_name', 'Farreach Electronic Co., Ltd.')
    ws['A1'].font = Font(bold=True, size=18)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 合并 A2:D2 - 标语
    ws.merge_cells('A2:D2')
    ws['A2'] = company.get('company_tagline', 'Premium Connectivity Solutions')
    ws['A2'].font = Font(size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # 空行
    ws.row_dimensions[3].height = 10
    
    # 合并 A4:D4 - QUOTATION 标题
    ws.merge_cells('A4:D4')
    ws['A4'] = 'Q U O T A T I O N'
    ws['A4'].font = Font(bold=True, size=24, color='0066CC')
    ws['A4'].alignment = Alignment(horizontal='center')
    
    # 空行
    ws.row_dimensions[5].height = 15
    
    # ============ 第 2 部分：报价单信息和客户信息（对称） ============
    # 左半部分：报价单信息
    ws['A6'] = 'Quotation No:'
    ws['A6'].font = Font(size=10, bold=True)
    ws['B6'] = quotation.get('quotation_no', 'QT-' + datetime.now().strftime('%Y%m%d-001'))
    
    ws['A7'] = 'Date:'
    ws['A7'].font = Font(size=10, bold=True)
    ws['B7'] = quotation.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    ws['A8'] = 'Valid Until:'
    ws['A8'].font = Font(size=10, bold=True)
    ws['B8'] = quotation.get('valid_until', '')
    
    # 右半部分：客户信息
    ws['C6'] = 'Customer:'
    ws['C6'].font = Font(size=10, bold=True)
    ws['D6'] = customer.get('company_name', customer.get('name', '_________________'))
    
    ws['C7'] = 'Attn:'
    ws['C7'].font = Font(size=10, bold=True)
    ws['D7'] = customer.get('contact', '_________________')
    
    ws['C8'] = 'Email:'
    ws['C8'].font = Font(size=10, bold=True)
    ws['D8'] = customer.get('email', '_________________')
    
    # 空行
    ws.row_dimensions[9].height = 15
    
    # ============ 第 3 部分：产品表格 ============
    # 表头（第 10 行）
    headers = ['Item', 'Description', 'Quantity', 'Unit Price', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.border = create_border()
    
    ws.row_dimensions[10].height = 20
    
    # 产品数据
    row = 11
    for idx, product in enumerate(products, start=1):
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=1).border = create_border()
        
        # Description + 规格 inline（节省空间）
        desc = product.get('description', '')
        spec = product.get('specification', '')
        if spec:
            specs = [s.strip() for s in spec.split(',') if s.strip()]
            desc += '\n' + ' | '.join(specs)
        
        cell = ws.cell(row=row, column=2, value=desc)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = create_border()
        
        ws.cell(row=row, column=3).value = product.get('quantity', 0)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).border = create_border()
        
        ws.cell(row=row, column=4).value = product.get('unit_price', product.get('unitPrice', 0))
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=4).border = create_border()
        
        # 金额公式
        ws.cell(row=row, column=5).value = f'=C{row}*D{row}'
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5).border = create_border()
        
        ws.row_dimensions[row].height = 40  # 从 60 降到 40
        row += 1
    
    last_product_row = row - 1
    
    # 空行（减少）
    ws.row_dimensions[row].height = 6
    row += 1
    
    # ============ 第 4 部分：总计（对称布局） ============
    # 合并左边单元格
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1).value = 'Subtotal:'
    ws.cell(row=row, column=1).font = Font(size=10)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
    
    ws.cell(row=row, column=4).value = f'=SUM(E11:E{last_product_row})'
    ws.cell(row=row, column=4).number_format = '$#,##0.00'
    ws.cell(row=row, column=4).font = Font(bold=True, size=10)
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=4).border = create_border()
    ws.cell(row=row, column=5).border = create_border()
    ws.row_dimensions[row].height = 18
    row += 1
    
    # Freight
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1).value = 'Freight:'
    ws.cell(row=row, column=1).font = Font(size=10)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
    
    freight = company.get('freight', 0)
    ws.cell(row=row, column=4).value = f'${freight:,.2f}' if freight > 0 else 'To be advised'
    ws.cell(row=row, column=4).font = Font(bold=True, size=10)
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=4).border = create_border()
    ws.cell(row=row, column=5).border = create_border()
    ws.row_dimensions[row].height = 18
    row += 1
    
    # Total
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1).value = 'TOTAL:'
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
    
    ws.cell(row=row, column=4).value = f'=SUM(D{row-2}:D{row-1})' if freight > 0 else f'=D{row-2}'
    ws.cell(row=row, column=4).number_format = '$#,##0.00'
    ws.cell(row=row, column=4).font = Font(bold=True, size=14, color='0066CC')
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=4).border = create_border()
    ws.cell(row=row, column=5).border = create_border()
    ws.row_dimensions[row].height = 25
    
    # 给总计区域加背景色
    for col in ['D', 'E']:
        for r in range(row-2, row+1):
            ws[f'{col}{r}'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    
    row += 1
    
    # ============ 第 5 部分：贸易条款（对称两列） ============
    ws['A{row}'.format(row=row)] = 'Payment Terms:'
    ws['A{row}'.format(row=row)].font = Font(size=10, bold=True)
    ws['B{row}'.format(row=row)] = company.get('payment_terms', 'T/T 30% deposit, 70% before shipment')
    ws['B{row}'.format(row=row)].font = Font(size=10)
    
    ws['C{row}'.format(row=row)] = 'Lead Time:'
    ws['C{row}'.format(row=row)].font = Font(size=10, bold=True)
    ws['D{row}'.format(row=row)] = company.get('lead_time', '15-20 days')
    ws['D{row}'.format(row=row)].font = Font(size=10)
    row += 1
    
    ws['A{row}'.format(row=row)] = 'Currency:'
    ws['A{row}'.format(row=row)].font = Font(size=10, bold=True)
    ws['B{row}'.format(row=row)] = company.get('currency', 'USD')
    ws['B{row}'.format(row=row)].font = Font(size=10)
    
    ws['C{row}'.format(row=row)] = 'Incoterms:'
    ws['C{row}'.format(row=row)].font = Font(size=10, bold=True)
    ws['D{row}'.format(row=row)] = 'FOB Shenzhen'
    ws['D{row}'.format(row=row)].font = Font(size=10)
    
    row += 2
    
    # ============ 第 6 部分：备注和条款 ============
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1).value = 'Remarks:'
    ws.cell(row=row, column=1).font = Font(size=10, bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    row += 1
    
    notes = company.get('notes', '1. Price based on current material cost.\n2. Valid for 30 days.')
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1).value = notes
    ws.cell(row=row, column=1).font = Font(size=9, color='666666')
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[row].height = 40
    
    row += 2
    
    # ============ 第 7 部分：银行信息 ============
    ws['A{row}'.format(row=row)] = 'Bank Details:'
    ws['A{row}'.format(row=row)].font = Font(size=10, bold=True)
    
    bank_info = company.get('bank_info', {})
    bank_data = [
        ('Beneficiary:', bank_info.get('beneficiary', 'Farreach Electronic Co., Ltd.')),
        ('Bank Name:', bank_info.get('bank_name', 'Standard Chartered Bank')),
        ('Account No:', bank_info.get('account_no', '1234 5678 9012')),
        ('SWIFT Code:', bank_info.get('swift_code', 'SCBLHKHH'))
    ]
    
    for i, (label, value) in enumerate(bank_data, start=1):
        ws.cell(row=row+i, column=1).value = label
        ws.cell(row=row+i, column=1).font = Font(size=9)
        ws.cell(row=row+i, column=2).value = value
        ws.cell(row=row+i, column=2).font = Font(size=9)
    
    row += len(bank_data) + 1
    
    # ============ 第 8 部分：签名 ============
    ws['C{row}'.format(row=row)] = 'Authorized Signature:'
    ws['C{row}'.format(row=row)].font = Font(size=10, bold=True)
    row += 2
    
    ws['C{row}'.format(row=row)] = '_________________________'
    ws['C{row}'.format(row=row)].font = Font(size=10)
    row += 1
    
    ws['C{row}'.format(row=row)] = 'Sales Manager'
    ws['C{row}'.format(row=row)].font = Font(size=9, italic=True, color='666666')
    
    # 保存
    wb.save(output_path)
    wb.close()
    
    return output_path

def main():
    parser = argparse.ArgumentParser(
        description='生成 Excel 报价单（传统工整风格）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  # 从 JSON 数据生成
  python3 generate_quotation_traditional.py --data quotation_data.json --output QT-20260314-001.xlsx
  
  # 快速测试
  python3 generate_quotation_traditional.py --output test.xlsx --quick-test
        '''
    )
    
    parser.add_argument('--data', '-d', help='报价数据 JSON 文件路径')
    parser.add_argument('--output', '-o', help='输出文件路径')
    parser.add_argument('--quick-test', action='store_true', help='使用测试数据快速生成')
    
    args = parser.parse_args()
    
    if args.output:
        data = None
        
        if args.data:
            with open(args.data, 'r', encoding='utf-8') as f:
                data = json.load(f)
        elif args.quick_test:
            data = {
                'company_name': 'Farreach Electronic Co., Ltd.',
                'company_tagline': 'Premium Connectivity Solutions',
                'customer': {
                    'company_name': 'Best Buy Electronics Inc.',
                    'contact': 'Michael Johnson',
                    'email': 'mjohnson@bestbuy-example.com'
                },
                'quotation': {
                    'quotation_no': 'QT-20260314-001',
                    'date': '2026-03-14',
                    'valid_until': '2026-04-13'
                },
                'products': [
                    {
                        'description': 'HDMI 2.1 Ultra High Speed Cable',
                        'specification': '8K@60Hz, 48Gbps, 2m',
                        'quantity': 500,
                        'unit_price': 8.50
                    },
                    {
                        'description': 'USB-C to USB-C Cable',
                        'specification': 'USB 4.0, 80Gbps, 100W PD, 1m',
                        'quantity': 1000,
                        'unit_price': 12.00
                    }
                ],
                'currency': 'USD',
                'payment_terms': 'T/T 30% deposit, 70% before shipment',
                'lead_time': '15-20 days',
                'freight': 350.00,
                'bank_info': {
                    'beneficiary': 'Farreach Electronic Co., Ltd.',
                    'bank_name': 'Standard Chartered Bank',
                    'account_no': '1234 5678 9012',
                    'swift_code': 'SCBLHKHH'
                }
            }
        
        if not args.data and not args.quick_test:
            print("❌ 请提供 --data 或 --quick-test", file=sys.stderr)
            sys.exit(1)
        
        output = create_quotation_excel(args.output, data)
        print(f"✅ Excel 报价单已生成：{output}")
        return
    
    parser.print_help()

if __name__ == '__main__':
    main()
