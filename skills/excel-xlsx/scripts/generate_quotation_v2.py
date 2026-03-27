#!/usr/bin/env python3
"""生成 Excel 报价单 v2 - 参考 HTML 结构，简洁专业"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter

# 颜色定义（与 HTML 一致）
COLORS = {
    'dark': '0F172A',      # slate-900 - 表头深色
    'primary': '0D9488',   # teal-600 - 强调色
    'light': 'F8FAFC',     # slate-50 - 浅灰背景
    'medium': '64748B',    # slate-500 - 次要文字
}

def create_border(style='thin'):
    """创建边框"""
    side = Side(style=style, color='000000')
    return Border(top=side, left=side, right=side, bottom=side)

def create_quotation_excel(output_path, data=None):
    """创建 Excel 报价单（参考 HTML 结构）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    
    # ============ 页面设置 ============
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    # 列宽（6 列）
    ws.column_dimensions['A'].width = 8    # No.
    ws.column_dimensions['B'].width = 50   # Description
    ws.column_dimensions['C'].width = 12   # Qty
    ws.column_dimensions['D'].width = 15   # Unit Price
    ws.column_dimensions['E'].width = 15   # Amount
    ws.column_dimensions['F'].width = 15   # 备用
    
    # ============ 数据准备 ============
    company = data or {}
    customer = company.get('customer', {})
    quotation = company.get('quotation', {})
    products = company.get('products', [])
    trade_terms = company.get('trade_terms', {})
    
    # ============ 第 1 部分：Header（公司和报价单信息） ============
    row = 1
    
    # 公司名称（A1）
    ws['A1'] = company.get('company_name', 'Farreach Electronic Co., Ltd.')
    ws['A1'].font = Font(bold=True, size=18, color='0F172A')
    
    # 公司标语（A2）
    ws['A2'] = company.get('company_tagline', 'Premium Connectivity Solutions')
    ws['A2'].font = Font(size=10, color='64748B', italic=True)
    
    # 联系信息（A4:A6）
    ws['A4'] = f"📧 {company.get('company_email', 'sale@farreach-electronic.com')}"
    ws['A4'].font = Font(size=9, color='64748B')
    
    ws['A5'] = f"🌐 {company.get('company_website', 'www.farreach-electronic.com')}"
    ws['A5'].font = Font(size=9, color='64748B')
    
    ws['A6'] = f"📍 {company.get('company_address', 'Zhuhai, Guangdong, China')}"
    ws['A6'].font = Font(size=9, color='64748B')
    
    # 报价单标题（E1）
    ws['E1'] = "QUOTATION"
    ws['E1'].font = Font(bold=True, size=20, color='0D9488')
    ws['E1'].alignment = Alignment(horizontal='right')
    
    # 报价单信息（E3:E5）
    ws['E3'] = "Quote No:"
    ws['E3'].font = Font(size=9, color='64748B')
    ws['F3'] = quotation.get('quotation_no', 'QT-' + datetime.now().strftime('%Y%m%d-001'))
    ws['F3'].font = Font(bold=True, size=9)
    
    ws['E4'] = "Date:"
    ws['E4'].font = Font(size=9, color='64748B')
    ws['F4'] = quotation.get('date', datetime.now().strftime('%Y-%m-%d'))
    ws['F4'].font = Font(bold=True, size=9)
    
    ws['E5'] = "Valid Until:"
    ws['E5'].font = Font(size=9, color='64748B')
    ws['F5'] = quotation.get('valid_until', '')
    ws['F5'].font = Font(bold=True, size=9)
    
    # 空行
    row = 7
    ws.row_dimensions[row].height = 15
    
    # ============ 第 2 部分：客户信息和贸易条款 ============
    row = 8
    
    # 客户信息标题（A8）
    ws['A8'] = "Prepared For"
    ws['A8'].font = Font(bold=True, size=10, color='64748B')
    ws['A8'].fill = PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid')
    
    # 客户信息（A9:A12）
    ws['A9'] = customer.get('company_name', '_________________')
    ws['A9'].font = Font(bold=True, size=12)
    
    ws['A10'] = f"Attn: {customer.get('contact', '_________________')}"
    ws['A10'].font = Font(size=10)
    
    ws['A11'] = customer.get('address', '_________________')
    ws['A11'].font = Font(size=10)
    
    ws['A12'] = customer.get('email', '_________________')
    ws['A12'].font = Font(size=10)
    
    # 贸易条款标题（D8）
    ws['D8'] = "Trade Terms"
    ws['D8'].font = Font(bold=True, size=10, color='64748B')
    ws['D8'].fill = PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid')
    
    # 贸易条款（D9:D12）
    ws['D9'] = "Incoterms:"
    ws['D9'].font = Font(size=9, color='64748B')
    ws['E9'] = trade_terms.get('incoterms', 'FOB Shenzhen')
    ws['E9'].font = Font(bold=True, size=9)
    
    ws['D10'] = "Currency:"
    ws['D10'].font = Font(size=9, color='64748B')
    ws['E10'] = company.get('currency', 'USD')
    ws['E10'].font = Font(bold=True, size=9)
    
    ws['D11'] = "Lead Time:"
    ws['D11'].font = Font(size=9, color='64748B')
    ws['E11'] = company.get('lead_time', '15-20 days')
    ws['E11'].font = Font(bold=True, size=9)
    
    ws['D12'] = "Payment Terms:"
    ws['D12'].font = Font(size=9, color='64748B')
    ws['E12'] = company.get('payment_terms', 'T/T 30% deposit, 70% before shipment')
    ws['E12'].font = Font(bold=True, size=9)
    
    # 空行
    row = 13
    ws.row_dimensions[row].height = 15
    
    # ============ 第 3 部分：产品表格 ============
    row = 14
    
    # 表头（A14:E14）
    headers = [
        ('No.', 'center'),
        ('Description & Specifications', 'left'),
        ('Qty', 'center'),
        ('Unit Price', 'right'),
        ('Amount', 'right')
    ]
    
    for col, (header_text, align) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header_text
        cell.font = Font(bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['dark'], end_color=COLORS['dark'], fill_type='solid')
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = create_border()
        ws.row_dimensions[row].height = 25
    
    # 产品数据
    row = 15
    for idx, product in enumerate(products, start=1):
        ws[f'A{row}'] = f'{idx:02d}'
        ws[f'A{row}'].font = Font(size=9, color='64748B')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].border = create_border()
        
        ws[f'B{row}'] = product.get('description', '')
        ws[f'B{row}'].font = Font(bold=True, size=9)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'B{row}'].border = create_border()
        
        # 规格作为第二行
        spec_text = product.get('specification', '')
        if spec_text:
            specs = [s.strip() for s in spec_text.split(',') if s.strip()]
            spec_list = '\n'.join(f'• {spec}' for spec in specs)
            ws[f'B{row}'].value = f"{product.get('description', '')}\n\n{spec_list}"
        
        ws[f'C{row}'] = product.get('quantity', 0)
        ws[f'C{row}'].font = Font(size=9)
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        ws[f'C{row}'].border = create_border()
        
        ws[f'D{row}'] = product.get('unit_price', 0)
        ws[f'D{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].font = Font(size=9)
        ws[f'D{row}'].alignment = Alignment(horizontal='right')
        ws[f'D{row}'].border = create_border()
        
        # 金额公式
        ws[f'E{row}'] = f'=C{row}*D{row}'
        ws[f'E{row}'].number_format = '$#,##0.00'
        ws[f'E{row}'].font = Font(bold=True, size=9)
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].border = create_border()
        
        ws.row_dimensions[row].height = 60
        row += 1
    
    # 最后一行行号
    last_product_row = row - 1
    
    # 空行
    ws.row_dimensions[row].height = 15
    row += 1
    
    # ============ 第 4 部分：总计（全宽） ============
    # 合并单元格 A-E
    ws.merge_cells(f'A{row}:E{row}')
    total_cell = ws[f'A{row}']
    total_cell.fill = PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid')
    ws.row_dimensions[row].height = 10
    row += 1
    
    # Subtotal
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Subtotal"
    ws[f'A{row}'].font = Font(size=10, color='64748B')
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    
    ws[f'D{row}'] = f'=SUM(E15:E{last_product_row})'
    ws[f'D{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].font = Font(bold=True, size=10)
    ws[f'D{row}'].alignment = Alignment(horizontal='right')
    ws.row_dimensions[row].height = 20
    row += 1
    
    # Freight
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Estimated Freight"
    ws[f'A{row}'].font = Font(size=10, color='64748B', italic=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    
    freight = company.get('freight', 0)
    ws[f'D{row}'] = f'${freight:,.2f}' if freight > 0 else 'To be advised'
    ws[f'D{row}'].font = Font(bold=True, size=10)
    ws[f'D{row}'].alignment = Alignment(horizontal='right')
    ws.row_dimensions[row].height = 20
    row += 1
    
    # Total
    ws.merge_cells(f'A{row}:C{row}')
    incoterms = trade_terms.get('incoterms', 'FOB')
    ws[f'A{row}'] = f"Total ({incoterms})"
    ws[f'A{row}'].font = Font(bold=True, size=14, color='0F172A')
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    
    ws[f'D{row}'] = f'=SUM(D{row-2}:D{row-1})' if freight > 0 else f'=D{row-2}'
    ws[f'D{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].font = Font(bold=True, size=16, color='0D9488')
    ws[f'D{row}'].alignment = Alignment(horizontal='right')
    ws.row_dimensions[row].height = 30
    
    # 添加边框
    for col in ['A', 'B', 'C', 'D', 'E']:
        for r in range(row-3, row+1):
            ws[f'{col}{r}'].border = create_border()
    
    # ============ 第 5 部分：条款和银行信息 ============
    row += 3
    
    # 条款标题（A{row}）
    ws[f'A{row}'] = "Terms & Conditions"
    ws[f'A{row}'].font = Font(bold=True, size=11, color='0F172A')
    
    row += 1
    terms = company.get('terms', [
        'Payment: 30% T/T deposit in advance, 70% balance before shipment.',
        'Packaging: Standard PE bag. Custom retail packaging available upon request.',
        'Warranty: 12 months against manufacturing defects.',
        'Certification: CE, FCC, RoHS, HDMI compliance guaranteed.'
    ])
    
    for i, term in enumerate(terms, start=1):
        ws[f'A{row}'] = f"{i}. {term}"
        ws[f'A{row}'].font = Font(size=9, color='64748B')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 15
        row += 1
    
    # 银行信息（D{row-4}）
    bank_row = row - 4
    ws[f'D{bank_row}'] = "Bank Details"
    ws[f'D{bank_row}'].font = Font(bold=True, size=11, color='0F172A')
    
    bank_row += 1
    bank_info = company.get('bank_info', {})
    bank_data = [
        ('Beneficiary:', bank_info.get('beneficiary', 'Farreach Electronic Co., Ltd.')),
        ('Bank Name:', bank_info.get('bank_name', 'Standard Chartered Bank')),
        ('Account No:', bank_info.get('account_no', '1234 5678 9012')),
        ('SWIFT Code:', bank_info.get('swift_code', 'SCBLHKHH'))
    ]
    
    for label, value in bank_data:
        ws[f'D{bank_row}'] = label
        ws[f'D{bank_row}'].font = Font(size=9, color='64748B')
        ws[f'E{bank_row}'] = value
        ws[f'E{bank_row}'].font = Font(size=9)
        ws.row_dimensions[bank_row].height = 15
        bank_row += 1
    
    # ============ 第 6 部分：签名区域（第二页） ============
    row += 3
    
    # 添加分页符
    ws.sheet_view.showGridLines = False
    
    ws[f'A{row}'] = "Authorized Signature:"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    
    row += 3
    ws[f'A{row}'] = "_________________________"
    ws[f'A{row}'].font = Font(size=11)
    
    row += 1
    ws[f'A{row}'] = "Sales Manager"
    ws[f'A{row}'].font = Font(size=10, italic=True, color='64748B')
    
    # 保存
    wb.save(output_path)
    wb.close()
    
    return output_path

def main():
    parser = argparse.ArgumentParser(
        description='生成 Excel 报价单 v2（参考 HTML 结构）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  # 从 JSON 数据生成
  python3 generate_quotation_v2.py --data quotation_data.json --output QT-20260314-001.xlsx
  
  # 快速测试
  python3 generate_quotation_v2.py --output test.xlsx --quick-test
        '''
    )
    
    parser.add_argument('--data', '-d', help='报价数据 JSON 文件路径')
    parser.add_argument('--output', '-o', help='输出文件路径')
    parser.add_argument('--quick-test', action='store_true', help='使用测试数据快速生成')
    
    args = parser.parse_args()
    
    if args.output:
        data = None
        
        if args.data:
            with open(args.data, 'r', encoding='utf-8') as f:
                data = json.load(f)
        elif args.quick_test:
            data = {
                'company_name': 'Farreach Electronic Co., Ltd.',
                'company_tagline': 'Premium Connectivity Solutions',
                'company_email': 'sale@farreach-electronic.com',
                'company_website': 'www.farreach-electronic.com',
                'company_address': 'Zhuhai, Guangdong, China',
                'customer': {
                    'company_name': 'Best Buy Electronics Inc.',
                    'contact': 'Michael Johnson',
                    'address': '7601 Penn Avenue South, Richfield, MN 55423, USA',
                    'email': 'mjohnson@bestbuy-example.com'
                },
                'quotation': {
                    'quotation_no': 'QT-20260314-001',
                    'date': '2026-03-14',
                    'valid_until': '2026-04-13'
                },
                'trade_terms': {
                    'incoterms': 'FOB Shenzhen'
                },
                'products': [
                    {
                        'description': 'HDMI 2.1 Ultra High Speed Cable',
                        'specification': '8K@60Hz, 4K@120Hz, 48Gbps, HDR, eARC, 2m',
                        'quantity': 500,
                        'unit_price': 8.50
                    },
                    {
                        'description': 'HDMI 2.1 Fiber Optical Cable (AOC)',
                        'specification': '8K@60Hz, 48Gbps, Active Optical, 10m',
                        'quantity': 200,
                        'unit_price': 25.00
                    },
                    {
                        'description': 'USB-C to USB-C Cable',
                        'specification': 'USB 4.0, 80Gbps, 100W PD 2.0, 1m',
                        'quantity': 1000,
                        'unit_price': 12.00
                    }
                ],
                'currency': 'USD',
                'lead_time': '15-20 days',
                'payment_terms': 'T/T 30% deposit, 70% before shipment',
                'freight': 350.00,
                'bank_info': {
                    'beneficiary': 'Farreach Electronic Co., Ltd.',
                    'bank_name': 'Standard Chartered Bank',
                    'account_no': '1234 5678 9012',
                    'swift_code': 'SCBLHKHH'
                }
            }
        
        if not args.data and not args.quick_test:
            print("❌ 请提供 --data 或 --quick-test", file=sys.stderr)
            sys.exit(1)
        
        output = create_quotation_excel(args.output, data)
        print(f"✅ Excel 报价单已生成：{output}")
        return
    
    parser.print_help()

if __name__ == '__main__':
    main()
