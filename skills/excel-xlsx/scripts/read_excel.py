#!/usr/bin/env python3
"""读取 Excel 文件并输出结构化数据 - 支持多 sheet、类型转换、日期处理"""

import sys
import json
import argparse
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, date

def excel_to_serial(dt):
    """Python datetime → Excel 序列号（1900 日期系统）"""
    if dt is None:
        return None
    
    # Excel 的 1900-01-01 是序列号 1
    # 注意：Excel 错误地把 1900 年当作闰年（序列号 60 = 1900-02-29，实际不存在）
    excel_epoch = date(1899, 12, 30)  # 修正 1900 闰年 bug
    
    if isinstance(dt, datetime):
        delta = dt - datetime.combine(excel_epoch, datetime.min.time())
    else:
        delta = dt - excel_epoch
    
    return delta.days + (dt.hour / 24 + dt.minute / 1440 + dt.second / 86400) if isinstance(dt, datetime) else delta.days

def serial_to_excel(serial):
    """Excel 序列号 → Python datetime（1900 日期系统）"""
    if serial is None:
        return None
    
    # Excel 的 1900-01-01 是序列号 1
    # 修正 1900 闰年 bug：序列号 60 之后需要减 1 天
    if serial < 60:
        base = date(1899, 12, 30)
    else:
        base = date(1899, 12, 31)  # 跳过不存在的 1900-02-29
    
    days = int(serial)
    fraction = serial - days
    
    result = date(base.year, base.month, base.day)
    from datetime import timedelta
    result = base + timedelta(days=days - 1)
    
    # 处理时间部分
    if fraction > 0:
        hours = int(fraction * 24)
        minutes = int((fraction * 24 - hours) * 60)
        seconds = int((fraction * 1440 - hours * 60 - minutes) * 60)
        return datetime(result.year, result.month, result.day, hours, minutes, seconds)
    
    return result

def cell_value(cell, workbook_date_system='1900'):
    """智能提取单元格值，处理日期、数字、文本类型"""
    value = cell.value
    
    if value is None:
        return None
    
    # 检查是否是日期类型
    if cell.is_date or isinstance(value, (datetime, date)):
        if isinstance(value, (int, float)):
            # 日期序列号
            return serial_to_excel(value)
        return value.isoformat() if hasattr(value, 'isoformat') else str(value)
    
    # 检查是否是公式
    if cell.data_type == 'f':
        # 返回公式字符串
        return f"={value}"
    
    # 数字类型
    if isinstance(value, (int, float)):
        # 检查是否是长数字（超过 15 位需要特殊处理）
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return value
    
    # 布尔类型
    if isinstance(value, bool):
        return value
    
    # 字符串类型
    return str(value)

def read_excel(path, sheet=None, output_format='json', verbose=False):
    """读取 Excel 文件
    
    Args:
        path: 文件路径（支持通配符）
        sheet: 指定 sheet 名称，None 表示所有 sheet
        output_format: 'json' | 'csv' | 'table'
        verbose: 是否显示详细信息
    """
    import glob
    
    files = glob.glob(path)
    if not files:
        print(f"❌ 找不到文件：{path}", file=sys.stderr)
        sys.exit(1)
    
    if verbose:
        print(f"📁 找到 {len(files)} 个文件\n")
    
    results = []
    
    for file_path in sorted(files):
        if verbose:
            print(f"📄 读取：{Path(file_path).name}")
        
        try:
            wb = load_workbook(filename=file_path, data_only=True)
            
            # 检查日期系统（1900 或 1904）
            date_system = '1900'
            if wb.properties and hasattr(wb.properties, 'date1904'):
                date_system = '1904' if wb.properties.date1904 else '1900'
            
            sheet_names = [sheet] if sheet else wb.sheetnames
            
            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    print(f"⚠️  Sheet '{sheet_name}' 不存在，跳过", file=sys.stderr)
                    continue
                
                ws = wb[sheet_name]
                
                # 读取数据
                data = []
                headers = None
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    # 跳过全空行
                    if all(cell is None for cell in row):
                        continue
                    
                    # 第一行作为表头
                    if row_idx == 1:
                        headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(row)]
                        continue
                    
                    # 数据行
                    row_data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers):
                            # 获取完整单元格对象以检查类型
                            cell_obj = ws.cell(row=row_idx, column=col_idx + 1)
                            row_data[headers[col_idx]] = cell_value(cell_obj, date_system)
                    
                    data.append(row_data)
                
                result = {
                    'file': Path(file_path).name,
                    'sheet': sheet_name,
                    'headers': headers,
                    'data': data,
                    'row_count': len(data)
                }
                
                results.append(result)
                
                if verbose:
                    print(f"  ✅ Sheet '{sheet_name}': {len(data)} 行数据")
            
            wb.close()
            
        except Exception as e:
            print(f"❌ 处理 {file_path} 失败：{e}", file=sys.stderr)
            if verbose:
                import traceback
                traceback.print_exc()
    
    return results

def format_output(results, output_format='json'):
    """格式化输出"""
    if output_format == 'json':
        print(json.dumps(results, indent=2, ensure_ascii=False, default=str))
    
    elif output_format == 'csv':
        import csv
        writer = csv.writer(sys.stdout)
        for result in results:
            if result['headers'] and result['data']:
                # 写入表头
                writer.writerow([f"[{result['file']}] {result['sheet']}"] + result['headers'])
                # 写入数据
                for row in result['data']:
                    writer.writerow([result['file']] + [row.get(h, '') for h in result['headers']])
    
    elif output_format == 'table':
        for result in results:
            print(f"\n{'='*60}")
            print(f"📄 {result['file']} | Sheet: {result['sheet']} | {result['row_count']} 行")
            print(f"{'='*60}")
            
            if result['headers']:
                # 打印表头
                print(" | ".join(str(h)[:20].ljust(20) for h in result['headers']))
                print("-" * 60)
                
                # 打印前 10 行数据
                for i, row in enumerate(result['data'][:10], start=1):
                    values = [str(row.get(h, ''))[:20].ljust(20) for h in result['headers']]
                    print(f"{i:3}. " + " | ".join(values))
                
                if len(result['data']) > 10:
                    print(f"... 还有 {len(result['data']) - 10} 行")
    
    else:
        print(f"❌ 未知输出格式：{output_format}", file=sys.stderr)
        sys.exit(1)

def main():
    parser = argparse.ArgumentParser(
        description='读取 Excel 文件并输出结构化数据',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  python3 read_excel.py "data/products.xlsx"
  python3 read_excel.py "data/*.xlsx" --sheet "产品列表"
  python3 read_excel.py "quotation.xlsx" --format table
  python3 read_excel.py "data.xlsx" -v
        '''
    )
    
    parser.add_argument('path', help='Excel 文件路径（支持通配符）')
    parser.add_argument('--sheet', '-s', help='指定 Sheet 名称，默认读取所有')
    parser.add_argument('--format', '-f', choices=['json', 'csv', 'table'], default='json',
                        help='输出格式（默认：json）')
    parser.add_argument('--verbose', '-v', action='store_true', help='显示详细信息')
    
    args = parser.parse_args()
    
    results = read_excel(args.path, sheet=args.sheet, output_format=args.format, verbose=args.verbose)
    format_output(results, args.format)

if __name__ == '__main__':
    main()
