#!/usr/bin/env python3
"""生成 Excel 报价单 - 支持模板填充、公式、格式、多 sheet

🔴 P0-REVISE: 集成数据验证（防止示例数据）
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# 🔴 P0: 导入验证模块
sys.path.insert(0, str(Path(__file__).parent.parent / 'quotation-workflow' / 'scripts'))
try:
    from quotation_schema import validate_quotation_data
    VALIDATION_AVAILABLE = True
except ImportError:
    VALIDATION_AVAILABLE = False
    print("⚠️  警告：quotation_schema 模块不可用，将跳过数据验证")

# 颜色定义
COLORS = {
    'header': '4472C4',      # 蓝色表头
    'title': 'D9E1F2',       # 浅蓝标题
    'accent': 'FFC000',      # 黄色强调
    'success': '70AD47',     # 绿色
    'warning': 'FFC000',     # 橙色警告
    'danger': 'C00000',      # 红色
    'light_gray': 'F2F2F2',  # 浅灰背景
}

def create_border():
    """创建标准边框"""
    thin = Side(style='thin', color='000000')
    return Border(top=thin, left=thin, right=thin, bottom=thin)

def apply_header_style(cell):
    """应用表头样式到单元格"""
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = create_border()

def apply_title_style(cell):
    """应用标题样式到单元格"""
    cell.font = Font(bold=True, size=16, color='1F4E78')
    cell.alignment = Alignment(horizontal='center', vertical='center')

def create_currency_format():
    """创建货币格式"""
    return '#,##0.00'

def create_date_format():
    """创建日期格式"""
    return 'YYYY-MM-DD'

def excel_serial_from_date(dt):
    """Python date → Excel 序列号"""
    if dt is None:
        return None
    
    from datetime import timedelta
    excel_epoch = date(1899, 12, 30)
    
    if isinstance(dt, datetime):
        delta = dt - datetime.combine(excel_epoch, datetime.min.time())
    else:
        delta = dt - excel_epoch
    
    return delta.days

def create_quotation_template(output_path, data=None):
    """创建标准报价单模板
    
    Args:
        output_path: 输出文件路径
        data: 报价数据字典
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "报价单"
    
    # ============ 表头区域 ============
    # 公司 Logo 和标题（第 1-3 行）
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "报价单 QUOTATION"
    apply_title_style(title_cell)
    ws.row_dimensions[1].height = 40
    
    # 公司信息（第 2 行）
    ws.merge_cells('A2:F2')
    ws['A2'].value = "Farreach Electronic Co., Ltd. | 珠海 + 越南双基地 | 18 年经验"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20
    
    # 联系信息（第 3 行）
    ws.merge_cells('A3:F3')
    ws['A3'].value = "📧 your-email | 🌐 www.farreach-electronic.com | 📱 +86-756-XXXXXXX"
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 15
    
    # 空行（第 4 行）
    ws.row_dimensions[4].height = 10
    
    # ============ 客户信息和报价单号（第 5-8 行） ============
    # 表头
    headers_info = ["客户信息 Customer", "", "报价单信息 Quotation", ""]
    for col, header in enumerate(headers_info, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        apply_header_style(cell)
    
    ws.merge_cells('A5:B5')
    ws.merge_cells('C5:D5')
    
    # 客户信息
    customer_data = data.get('customer', {}) if data else {}
    ws['A6'].value = "公司名称 Company:"
    ws['B6'].value = customer_data.get('company_name', customer_data.get('name', '_________________'))
    
    ws['A7'].value = "联系人 Contact:"
    ws['B7'].value = customer_data.get('contact', '_________________')
    
    ws['A8'].value = "邮箱 Email:"
    ws['B8'].value = customer_data.get('email', '_________________')
    
    # 报价单信息
    quotation_data = data.get('quotation', {}) if data else {}
    ws['C6'].value = "报价单号 Quotation No:"
    ws['D6'].value = quotation_data.get('quotation_no', 'QT-' + datetime.now().strftime('%Y%m%d-001'))
    
    ws['C7'].value = "报价日期 Date:"
    ws['D7'].value = quotation_data.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    ws['C8'].value = "有效期至 Valid Until:"
    ws['D8'].value = quotation_data.get('valid_until', '')
    
    # 设置列宽（优化后：序号列调窄，产品描述加宽）
    ws.column_dimensions['A'].width = 8    # 序号 - 原来 25 太宽
    ws.column_dimensions['B'].width = 40   # 产品描述 - 加宽
    ws.column_dimensions['C'].width = 35   # 规格型号 - 加宽
    ws.column_dimensions['D'].width = 12   # 数量
    ws.column_dimensions['E'].width = 14   # 单价
    ws.column_dimensions['F'].width = 16   # 金额
    
    # ============ 产品列表表头（第 10 行） ============
    row = 10
    product_headers = [
        "序号\nNo.",
        "产品描述\nProduct Description",
        "规格型号\nSpecification",
        "数量\nQuantity",
        "单价\nUnit Price (USD)",
        "金额\nAmount (USD)"
    ]
    
    for col, header in enumerate(product_headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_header_style(cell)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    ws.row_dimensions[row].height = 40
    
    # ============ 产品数据（第 11 行开始） ============
    products = data.get('products', []) if data else []
    
    for idx, product in enumerate(products, start=1):
        row = 10 + idx
        
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = product.get('description', '')
        ws.cell(row=row, column=3).value = product.get('specification', '')
        ws.cell(row=row, column=4).value = product.get('quantity', 0)
        ws.cell(row=row, column=5).value = product.get('unit_price', product.get('unitPrice', 0))
        
        # 金额公式
        ws.cell(row=row, column=6).value = f"=D{row}*E{row}"
        ws.cell(row=row, column=6).number_format = create_currency_format()
        
        # 应用样式
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = create_border()
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        ws.row_dimensions[row].height = 30
    
    # ============ 汇总区域 ============
    last_product_row = 10 + len(products)
    row = last_product_row + 2
    
    # 币别
    ws['A{row}'.format(row=row)].value = "币别 Currency:"
    ws['B{row}'.format(row=row)].value = (data.get('currency', 'USD') if data else 'USD')
    
    # 付款条款
    ws['C{row}'.format(row=row)].value = "付款条款 Payment Terms:"
    ws['D{row}'.format(row=row)].value = (data.get('payment_terms', 'T/T 30% deposit, 70% before shipment') if data else 'T/T')
    ws.merge_cells('D{row}:F{row}'.format(row=row))
    
    row += 1
    
    # 交货期
    ws['A{row}'.format(row=row)].value = "交货期 Lead Time:"
    ws['B{row}'.format(row=row)].value = (data.get('lead_time', '15-20 days after deposit') if data else '15-20 days')
    
    # 小计
    ws['D{row}'.format(row=row)].value = "小计 Subtotal:"
    ws['E{row}'.format(row=row)].value = f"=SUM(F11:F{last_product_row})"
    ws['E{row}'.format(row=row)].number_format = create_currency_format()
    
    row += 1
    
    # 运费
    ws['D{row}'.format(row=row)].value = "运费 Freight:"
    ws['E{row}'.format(row=row)].value = (data.get('freight', 0) if data else 0)
    ws['E{row}'.format(row=row)].number_format = create_currency_format()
    
    row += 1
    
    # 税费
    ws['D{row}'.format(row=row)].value = "税费 Tax:"
    ws['E{row}'.format(row=row)].value = (data.get('tax', 0) if data else 0)
    ws['E{row}'.format(row=row)].number_format = create_currency_format()
    
    row += 1
    
    # 总计
    ws['D{row}'.format(row=row)].value = "总计 Total:"
    ws['E{row}'.format(row=row)].value = f"=SUM(E{row-3}:E{row-1})"
    ws['E{row}'.format(row=row)].number_format = create_currency_format()
    
    # 总计样式
    for col in ['D', 'E']:
        cell = ws[f'{col}{row}']
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color=COLORS['title'], end_color=COLORS['title'], fill_type='solid')
        cell.border = create_border()
    
    # ============ 备注和条款（底部） ============
    row += 3
    
    ws.merge_cells(f'A{row}:F{row}')
    notes_cell = ws[f'A{row}']
    notes_cell.value = "备注 Remarks:\n" + (data.get('notes', '1. 以上价格基于当前原材料成本，如有变动将另行通知。\n2. 最终价格以确认为准。') if data else '1. 以上价格基于当前原材料成本\n2. 最终价格以确认为准')
    notes_cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 60
    
    row += 1
    
    ws.merge_cells(f'A{row}:F{row}')
    terms_cell = ws[f'A{row}']
    terms_cell.value = "条款 Terms & Conditions:\n1. 报价有效期：30 天\n2. 付款方式：T/T 或 L/C\n3. 包装：标准出口包装\n4. 运输：FOB Shenzhen 或 CIF"
    terms_cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 60
    
    # ============ 打印设置 ============
    # 设置打印区域（A1 到 F{row}）
    ws.print_area = f'A1:F{row}'
    
    # 设置所有列在一页宽度，高度自适应
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = None
    
    # 设置纸张为 A4 横向 (9 = A4)
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'landscape'  # 横向打印
    
    # 设置页边距（英寸）
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    
    # ============ 签名区域（第二页） ============
    # 留出空白行用于分页
    row += 5
    
    ws['A{row}'.format(row=row)].value = "授权签名 Authorized Signature:"
    ws['A{row}'.format(row=row)].font = Font(bold=True)
    
    row += 2
    ws['A{row}'.format(row=row)].value = "_________________________"
    
    row += 1
    ws['A{row}'.format(row=row)].value = "Sales Manager"
    ws['A{row}'.format(row=row)].font = Font(italic=True)
    
    # 保存文件
    wb.save(output_path)
    wb.close()
    
    return output_path

def generate_quotation_from_template(template_path, output_path, data):
    """从模板生成报价单（填充数据）"""
    # TODO: 实现模板填充逻辑
    pass

def main():
    parser = argparse.ArgumentParser(
        description='生成 Excel 报价单',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  # 创建空白模板
  python3 generate_quotation.py --template output/quotation_template.xlsx
  
  # 从 JSON 数据生成报价单
  python3 generate_quotation.py --data quotation_data.json --output output/QT-20260314-001.xlsx
  
  # 从命令行数据生成（快速测试）
  python3 generate_quotation.py --output test.xlsx --quick-test
        '''
    )
    
    parser.add_argument('--template', '-t', help='创建空白模板到指定路径')
    parser.add_argument('--data', '-d', help='报价数据 JSON 文件路径')
    parser.add_argument('--output', '-o', help='输出文件路径')
    parser.add_argument('--quick-test', action='store_true', help='使用测试数据快速生成')
    
    args = parser.parse_args()
    
    # 创建空白模板
    if args.template:
        output = create_quotation_template(args.template)
        print(f"✅ 模板已创建：{output}")
        return
    
    # 从数据生成
    if args.output:
        data = None
        
        if args.data:
            with open(args.data, 'r', encoding='utf-8') as f:
                data = json.load(f)
        elif args.quick_test:
            data = {
                'customer': {
                    'company_name': 'Test Customer Inc.',
                    'contact': 'John Doe',
                    'email': 'john@test.com'
                },
                'quotation': {
                    'quotation_no': 'QT-20260314-001',
                    'date': '2026-03-14',
                    'valid_until': '2026-04-13'
                },
                'products': [
                    {
                        'description': 'HDMI 2.1 Ultra High Speed Cable',
                        'specification': '8K@60Hz, 48Gbps, 2m',
                        'quantity': 500,
                        'unit_price': 8.50
                    },
                    {
                        'description': 'USB-C to USB-C Cable',
                        'specification': 'USB 4.0, 80Gbps, 100W PD, 1m',
                        'quantity': 1000,
                        'unit_price': 12.00
                    }
                ],
                'currency': 'USD',
                'payment_terms': 'T/T 30% deposit, 70% before shipment',
                'lead_time': '15-20 days after deposit',
                'freight': 150.00,
                'tax': 0,
                'notes': '1. 以上价格基于当前原材料成本\n2. 最终价格以确认为准'
            }
        
        if not args.data and not args.quick_test:
            print("❌ 请提供 --data 或 --quick-test", file=sys.stderr)
            sys.exit(1)
        
        # 🔴 P0: 数据验证（强制，无交互确认）
        if VALIDATION_AVAILABLE and not args.quick_test:
            print("🔍 验证报价单数据...")
            valid, errors = validate_quotation_data(data)
            
            if not valid:
                print("❌ 数据验证失败，Excel 报价单生成已终止:")
                for i, err in enumerate(errors, start=1):
                    print(f"  {i}. {err}")
                print()
                print("请检查数据文件，确保使用真实客户信息。")
                print("如需要测试，请使用 --quick-test 参数")
                sys.exit(1)
            
            print("✅ 数据验证通过")
            print()
        
        output = create_quotation_template(args.output, data)
        print(f"✅ Excel 报价单已生成：{output}")
        return
    
    # 没有参数时显示帮助
    parser.print_help()

if __name__ == '__main__':
    main()
