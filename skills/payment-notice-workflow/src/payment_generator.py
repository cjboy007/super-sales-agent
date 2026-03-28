#!/usr/bin/env python3
"""
收款通知生成器
"""

from dataclasses import dataclass
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os


@dataclass
class GenerationResult:
    success: bool
    output_path: Optional[str] = None
    error: Optional[str] = None


def generate_payment_notice_excel(data: dict, output_path: str) -> GenerationResult:
    """生成收款通知 Excel 文件"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Payment Notice'
        
        # 标题
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = 'PAYMENT NOTICE'
        title.font = Font(bold=True, size=16)
        title.alignment = Alignment(horizontal='center')
        
        # 收款通知信息
        notice = data.get('notice', {})
        ws['A3'] = f"Notice No: {notice.get('notice_no', 'N/A')}"
        ws['B3'] = f"Date: {notice.get('date', 'N/A')}"
        ws['C3'] = f"Due Date: {notice.get('due_date', 'N/A')}"
        
        # 客户信息
        customer = data.get('customer', {})
        ws['A5'] = 'To:'
        ws['A6'] = customer.get('company_name', '')
        ws['A7'] = customer.get('address', '')
        
        # 参考 PI
        reference = data.get('reference', {})
        ws['A9'] = f"Reference PI: {reference.get('pi_no', 'N/A')}"
        
        # 付款明细
        payment = data.get('payment', {})
        ws['A11'] = 'Payment Details:'
        ws['A12'] = 'Total Amount:'
        ws['B12'] = f"{payment.get('currency', 'USD')} {payment.get('total_amount', 0):.2f}"
        ws['A13'] = 'Deposit Paid:'
        ws['B13'] = f"{payment.get('currency', 'USD')} {payment.get('deposit_amount', 0):.2f}"
        ws['A14'] = 'Balance Due:'
        ws['B14'] = f"{payment.get('currency', 'USD')} {payment.get('balance_due', 0):.2f}"
        
        # 银行信息
        bank = data.get('bank', {})
        ws['A16'] = 'Bank Information:'
        ws['A17'] = f"Beneficiary: {bank.get('beneficiary', 'N/A')}"
        ws['A18'] = f"Bank Name: {bank.get('bank_name', 'N/A')}"
        ws['A19'] = f"Account No: {bank.get('account_no', 'N/A')}"
        ws['A20'] = f"SWIFT Code: {bank.get('swift_code', 'N/A')}"
        
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        
        wb.save(output_path)
        return GenerationResult(success=True, output_path=output_path)
    except Exception as e:
        return GenerationResult(success=False, error=str(e))
