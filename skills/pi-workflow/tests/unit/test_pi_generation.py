#!/usr/bin/env python3
"""
PI 生成测试
"""

import pytest
import json
import os
import sys
import subprocess

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'src'))
from pi_generator import generate_pi_excel

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), 'fixtures')
os.makedirs(FIXTURES_DIR, exist_ok=True)


def load_fixture(filename):
    filepath = os.path.join(FIXTURES_DIR, filename)
    if not os.path.exists(filepath):
        # 创建默认 fixture
        data = {
            "customer": {
                "company_name": "Test Customer Ltd",
                "contact": "John Doe",
                "email": "john@testcustomer.com",
                "phone": "+1-555-123-4567",
                "address": "123 Business St, City, Country",
                "country": "United States"
            },
            "products": [
                {"description": "HDMI Cable", "specification": "2m", "quantity": 100, "unit_price": 5.00}
            ],
            "pi": {"pi_no": "PI-20260328-001", "date": "2026-03-28", "valid_until": "2026-04-27"},
            "terms": {"payment": "T/T 30% deposit, 70% before shipment"}
        }
        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2)
        return data
    with open(filepath, 'r') as f:
        return json.load(f)


def test_pi_excel_created():
    """PI 应该能生成 Excel 文件"""
    data = load_fixture('pi_customer.json')
    output_path = os.path.join(os.path.dirname(__file__), 'output', 'test-pi.xlsx')
    
    result = generate_pi_excel(data, output_path)
    
    assert os.path.exists(output_path), f"PI 文件应该被创建：{output_path}"
    assert result.success, f"生成应该成功：{result.error}"


def find_cell_value(ws, value):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell.coordinate
    return None


def test_pi_excel_uses_external_pdf_source_layout():
    """对外 PI Excel 应该使用与 PDF 对应的完整源文件版式"""
    data = load_fixture('pi_customer.json')
    output_path = os.path.join(os.path.dirname(__file__), 'output', 'test-pi-layout.xlsx')

    result = generate_pi_excel(data, output_path)
    assert result.success, f"生成应该成功：{result.error}"

    wb = load_workbook(output_path, data_only=False)
    ws = wb.active

    assert ws['D1'].value == 'PROFORMA INVOICE'
    assert find_cell_value(ws, 'BILL TO:') is not None
    assert find_cell_value(ws, 'GOODS DESCRIPTION:') is not None
    assert find_cell_value(ws, 'PAYMENT TERMS & CONDITIONS:') is not None
    assert find_cell_value(ws, 'BANK DETAILS FOR PAYMENT:') is not None
    goods_cell = find_cell_value(ws, 'GOODS DESCRIPTION:')
    assert goods_cell is not None
    header_row = coordinate_to_tuple(goods_cell)[0] + 1
    assert [ws.cell(row=header_row, column=col).value for col in range(1, 6)] == [
        'No.', 'Description & Specifications', 'Qty', 'Unit Price', 'Amount'
    ]
    assert ws.freeze_panes == f'A{header_row + 1}'
    assert ws.print_area, "Excel 源文件必须设置打印区域，才能稳定导出 PDF"
    assert ws.page_setup.fitToWidth == 1


def test_pi_script_generates_matching_excel_source(tmp_path):
    """生成对外 HTML/PDF 时应同步生成同名 Excel 源文件"""
    fixture_path = os.path.join(FIXTURES_DIR, 'pi_customer.json')
    output_html = tmp_path / 'PI-20260328-001.html'
    script_path = os.path.join(os.path.dirname(__file__), '..', '..', 'scripts', 'generate_pi.py')

    completed = subprocess.run(
        [sys.executable, script_path, '--data', fixture_path, '--output', str(output_html)],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr or completed.stdout
    assert output_html.exists()
    assert output_html.with_suffix('.xlsx').exists()


def test_pi_number_format():
    """PI 编号格式应该是 PI-YYYYMMDD-XXX"""
    import re
    data = load_fixture('pi_customer.json')
    pi_no = data['pi']['pi_no']
    assert re.match(r'^PI-\d{8}-\d{3}$', pi_no), f"PI 编号格式错误：{pi_no}"


def test_pi_deposit_calculation():
    """PI 定金应该是 30%"""
    data = load_fixture('pi_customer.json')
    total = sum(p['quantity'] * p['unit_price'] for p in data['products'])
    deposit = total * 0.30
    assert deposit > 0
