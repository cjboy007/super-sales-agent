#!/usr/bin/env python3
"""
PI Excel 源文件生成器。

这份 Excel 是对外 Proforma Invoice PDF 的可编辑源文件，结构应与
scripts/generate_pi.py 生成的 HTML/PDF 保持一致。
"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass
class GenerationResult:
    success: bool
    output_path: Optional[str] = None
    error: Optional[str] = None


COLORS = {
    "dark": "111827",
    "light": "F3F4F6",
    "bank": "F9FAFB",
    "paper": "FFFFFF",
}


def border(style="thin"):
    side = Side(style=style, color="000000")
    return Border(top=side, left=side, right=side, bottom=side)


def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def as_money(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def as_date(value):
    return str(value or datetime.now().strftime("%Y-%m-%d")).replace("-", ".")


def money_format(currency: str):
    symbol = "$" if currency == "USD" else f'"{currency} "'
    return f'{symbol}#,##0.00'


def write_label_value(ws, row, label, value, label_col=1, value_col=2, width=5):
    ws.cell(row=row, column=label_col, value=label)
    ws.cell(row=row, column=value_col, value=value)
    ws.cell(row=row, column=label_col).font = Font(bold=True)
    ws.cell(row=row, column=label_col).alignment = Alignment(vertical="top")
    ws.cell(row=row, column=value_col).alignment = Alignment(wrap_text=True, vertical="top")
    if width > 2:
        ws.merge_cells(
            start_row=row,
            start_column=value_col,
            end_row=row,
            end_column=label_col + width - 1,
        )
    for col in range(label_col, label_col + width):
        ws.cell(row=row, column=col).border = border()


def section_title(ws, row, title, last_col=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=11, color="000000")
    cell.fill = fill(COLORS["light"])
    cell.alignment = Alignment(vertical="center")
    cell.border = border()
    ws.row_dimensions[row].height = 22


def style_table_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = fill(COLORS["dark"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()


def style_table_cell(cell, horizontal="left", bold=False, number_format=None):
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=True)
    cell.border = border()
    if number_format:
        cell.number_format = number_format


def generate_pi_excel(data: dict, output_path: str) -> GenerationResult:
    """生成与对外 PDF 版式匹配的 PI Excel 源文件"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Proforma Invoice"
        ws.sheet_view.showGridLines = False

        for column, width in {
            "A": 8,
            "B": 44,
            "C": 12,
            "D": 16,
            "E": 18,
        }.items():
            ws.column_dimensions[column].width = width

        company_name = data.get("company_name", "FARREACH ELECTRONIC CO LIMITED")
        company_address = data.get(
            "company_address",
            "No. 6, Chuangye Road East, Shuanglinpian, Liangang Industrial Zone, Zhuhai, China",
        )
        company_phone = data.get("company_phone", "YOUR-PHONE")
        company_fax = data.get("company_fax", "YOUR-FAX")
        company_website = data.get("company_website", "www.your-domain.com")
        company_email = data.get("company_email", "your-email")

        customer = data.get("customer", {})
        customer_name = customer.get("company_name", customer.get("name", "_________________"))
        customer_contact = customer.get("contact", customer.get("contact_name", "Procurement Manager"))
        customer_address = customer.get("address", "_________________")
        customer_email = customer.get("email", "")
        customer_phone = customer.get("phone", "")

        pi = data.get("pi", {})
        pi_no = pi.get("pi_no", data.get("piNo", "PI-" + datetime.now().strftime("%Y%m%d-001")))
        pi_date = as_date(pi.get("date", data.get("date")))
        valid_until = str(pi.get("valid_until", data.get("validUntil", ""))).replace("-", "/")

        trade_terms = data.get("trade_terms", {})
        terms = data.get("terms", {})
        incoterms = trade_terms.get("incoterms", terms.get("incoterms", "FOB Shenzhen"))
        currency = data.get("currency", terms.get("currency", "USD"))
        delivery = trade_terms.get("delivery", terms.get("delivery", "15-20 days"))

        products = data.get("products", [])
        freight = as_money(data.get("freight", 0))
        tax = as_money(data.get("tax", 0))
        subtotal = sum(as_money(p.get("quantity")) * as_money(p.get("unit_price", p.get("unitPrice", 0))) for p in products)
        total = subtotal + freight + tax

        bank_info = data.get("bank_info", {
            "beneficiary": "FARREACH ELECTRONIC CO LIMITED",
            "bank_name": "HSBC Hong Kong",
            "account_no": "411-758097-838",
            "swift_code": "HSBCHKHHHKH",
            "bank_address": "No.1 Queen's Road Central,Central, Hong Kong",
        })

        payment_terms = terms.get("payment", "T/T 30% deposit, 70% before shipment")
        packaging = terms.get("packaging", "Standard export packaging")

        ws.merge_cells("A1:C3")
        ws["A1"] = company_name
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(vertical="top", wrap_text=True)

        ws.merge_cells("A4:C7")
        ws["A4"] = (
            f"Address: {company_address}\n"
            f"Tel: {company_phone} | Fax: {company_fax}\n"
            f"Email: {company_email}\n"
            f"Website: {company_website}"
        )
        ws["A4"].font = Font(size=10)
        ws["A4"].alignment = Alignment(vertical="top", wrap_text=True)

        ws.merge_cells("D1:E3")
        ws["D1"] = "PROFORMA INVOICE"
        ws["D1"].font = Font(bold=True, size=18)
        ws["D1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(1, 4):
            for col in range(4, 6):
                ws.cell(row=row, column=col).border = border("medium")

        for row in range(1, 8):
            ws.row_dimensions[row].height = 22

        row = 9
        write_label_value(ws, row, "PI No:", pi_no, 1, 2, 3)
        write_label_value(ws, row, "Date:", pi_date, 4, 5, 2)
        row += 1
        write_label_value(ws, row, "Valid Until:", valid_until, 1, 2, 3)
        write_label_value(ws, row, "Currency:", currency, 4, 5, 2)

        row += 2
        section_title(ws, row, "BILL TO:")
        row += 1
        write_label_value(ws, row, "Company:", customer_name)
        row += 1
        write_label_value(ws, row, "Attn:", customer_contact)
        row += 1
        write_label_value(ws, row, "Address:", customer_address)
        ws.row_dimensions[row].height = 34
        row += 1
        if customer_email:
            write_label_value(ws, row, "Email:", customer_email)
            row += 1
        if customer_phone:
            write_label_value(ws, row, "Phone:", customer_phone)
            row += 1

        row += 1
        section_title(ws, row, "GOODS DESCRIPTION:")
        row += 1
        header_row = row
        headers = ["No.", "Description & Specifications", "Qty", "Unit Price", "Amount"]
        for col, header in enumerate(headers, start=1):
            style_table_header(ws.cell(row=header_row, column=col, value=header))
        ws.row_dimensions[header_row].height = 26

        first_product_row = header_row + 1
        row = first_product_row
        for idx, product in enumerate(products, start=1):
            description = product.get("description", "")
            specification = product.get("specification", "")
            quantity = as_money(product.get("quantity"))
            unit_price = as_money(product.get("unit_price", product.get("unitPrice", 0)))

            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=f"{description}\n{specification}".strip())
            ws.cell(row=row, column=3, value=quantity)
            ws.cell(row=row, column=4, value=unit_price)
            ws.cell(row=row, column=5, value=f"=C{row}*D{row}")

            style_table_cell(ws.cell(row=row, column=1), "center")
            style_table_cell(ws.cell(row=row, column=2))
            style_table_cell(ws.cell(row=row, column=3), "center")
            style_table_cell(ws.cell(row=row, column=4), "right", number_format=money_format(currency))
            style_table_cell(ws.cell(row=row, column=5), "right", bold=True, number_format=money_format(currency))
            ws.row_dimensions[row].height = 48
            row += 1

        last_product_row = row - 1
        if not products:
            row += 1

        subtotal_row = row + 1
        ws.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=4)
        ws.cell(row=subtotal_row, column=1, value="Subtotal:").alignment = Alignment(horizontal="right")
        ws.cell(row=subtotal_row, column=1).font = Font(bold=True)
        ws.cell(row=subtotal_row, column=5, value=f"=SUM(E{first_product_row}:E{row - 1})" if products else subtotal)

        freight_row = subtotal_row + 1
        ws.merge_cells(start_row=freight_row, start_column=1, end_row=freight_row, end_column=4)
        ws.cell(row=freight_row, column=1, value="Freight:").alignment = Alignment(horizontal="right")
        ws.cell(row=freight_row, column=1).font = Font(bold=True)
        ws.cell(row=freight_row, column=5, value=freight)

        tax_row = freight_row + 1
        ws.merge_cells(start_row=tax_row, start_column=1, end_row=tax_row, end_column=4)
        ws.cell(row=tax_row, column=1, value="Tax:").alignment = Alignment(horizontal="right")
        ws.cell(row=tax_row, column=1).font = Font(bold=True)
        ws.cell(row=tax_row, column=5, value=tax)

        total_row = tax_row + 1
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        ws.cell(row=total_row, column=1, value="TOTAL:").alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=total_row, column=5, value=f"=SUM(E{subtotal_row}:E{tax_row})" if products else total)
        ws.cell(row=total_row, column=5).font = Font(bold=True, size=12)

        for total_section_row in range(subtotal_row, total_row + 1):
            for col in range(1, 6):
                ws.cell(row=total_section_row, column=col).border = border()
                ws.cell(row=total_section_row, column=col).fill = fill(COLORS["light"]) if total_section_row == total_row else fill(COLORS["paper"])
            ws.cell(row=total_section_row, column=5).number_format = money_format(currency)
            ws.cell(row=total_section_row, column=5).alignment = Alignment(horizontal="right")

        row = total_row + 2
        section_title(ws, row, "PAYMENT TERMS & CONDITIONS:")
        row += 1
        write_label_value(ws, row, "Payment Method:", payment_terms)
        row += 1
        write_label_value(ws, row, "Delivery Time:", delivery)
        row += 1
        write_label_value(ws, row, "Port of Loading:", incoterms)
        row += 1
        write_label_value(ws, row, "Packaging:", packaging)

        row += 2
        section_title(ws, row, "BANK DETAILS FOR PAYMENT:")
        row += 1
        bank_rows = [
            ("Beneficiary:", bank_info.get("beneficiary", "")),
            ("Bank Name:", bank_info.get("bank_name", "")),
            ("Account No:", bank_info.get("account_no", "")),
            ("SWIFT Code:", bank_info.get("swift_code", "")),
            ("Bank Address:", bank_info.get("bank_address", "")),
        ]
        for label, value in bank_rows:
            write_label_value(ws, row, label, value)
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill(COLORS["bank"])
            row += 1

        row += 1
        section_title(ws, row, "TERMS & CONDITIONS:")
        row += 1
        write_label_value(ws, row, "Payment:", payment_terms)
        row += 1
        write_label_value(ws, row, "Delivery:", delivery)
        row += 1
        write_label_value(ws, row, "Packaging:", packaging)

        last_row = row
        for row_idx in range(1, last_row + 1):
            ws.row_dimensions[row_idx].height = ws.row_dimensions[row_idx].height or 22

        ws.freeze_panes = f"A{header_row + 1}"
        ws.print_area = f"A1:E{last_row}"
        ws.page_setup.paperSize = 9
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        ws.page_margins.left = 0.35
        ws.page_margins.right = 0.35
        ws.page_margins.top = 0.45
        ws.page_margins.bottom = 0.45
        ws.page_margins.header = 0
        ws.page_margins.footer = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.auto_filter.ref = f"A{header_row}:E{max(header_row, last_product_row)}"

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        return GenerationResult(success=True, output_path=str(output))
    except Exception as e:
        return GenerationResult(success=False, error=str(e))
