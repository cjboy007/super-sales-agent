#!/usr/bin/env python3
"""
样品单生成测试
"""

import pytest
import json
import os
import sys
import re
import subprocess

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'src'))
from sample_generator import generate_sample_excel

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), 'fixtures')
os.makedirs(FIXTURES_DIR, exist_ok=True)


def load_fixture(filename):
    filepath = os.path.join(FIXTURES_DIR, filename)
    if not os.path.exists(filepath):
        data = {
            "customer": {
                "company_name": "Test Customer Inc",
                "contact": "Mike Chen",
                "email": "mike@testcustomer.com",
                "phone": "+1-408-555-1234",
                "address": "456 Tech Drive, San Jose, CA",
                "country": "United States"
            },
            "products": [
                {"description": "HDMI Sample", "specification": "1m", "quantity": 5, "unit_price": 0.00}
            ],
            "sample": {"sample_no": "SPL-20260328-001", "date": "2026-03-28", "purpose": "Testing"},
            "shipping_address": {
                "company_name": "Test Customer Inc",
                "contact": "Mike Chen",
                "address": "456 Tech Drive, San Jose, CA"
            },
            "shipping": {"method": "DHL", "freight_collect": True}
        }
        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2)
        return data
    with open(filepath, 'r') as f:
        return json.load(f)


def test_sample_excel_created():
    """样品单应该能生成 Excel 文件"""
    data = load_fixture('sample_customer.json')
    output_path = os.path.join(os.path.dirname(__file__), 'output', 'test-sample.xlsx')
    
    result = generate_sample_excel(data, output_path)
    
    assert os.path.exists(output_path), f"样品单文件应该被创建：{output_path}"
    assert result.success, f"生成应该成功：{result.error}"


def find_cell_value(ws, value):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell.coordinate
    return None


def find_cell_containing(ws, value):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and value in cell.value:
                return cell.coordinate
    return None


def test_sample_excel_uses_external_pdf_source_layout():
    """对外样品单 Excel 应该使用与 PDF 对应的完整源文件版式"""
    data = load_fixture('sample_customer.json')
    output_path = os.path.join(os.path.dirname(__file__), 'output', 'test-sample-layout.xlsx')

    result = generate_sample_excel(data, output_path)
    assert result.success, f"生成应该成功：{result.error}"

    wb = load_workbook(output_path, data_only=False)
    ws = wb.active

    assert ws['E1'].value == 'SAMPLE REQUEST'
    assert find_cell_value(ws, 'CUSTOMER INFORMATION:') is not None
    assert find_cell_value(ws, 'SHIPPING ADDRESS:') is not None
    assert find_cell_value(ws, 'SAMPLE ITEMS:') is not None
    sample_items_cell = find_cell_value(ws, 'SAMPLE ITEMS:')
    assert sample_items_cell is not None
    header_row = coordinate_to_tuple(sample_items_cell)[0] + 1
    assert [ws.cell(row=header_row, column=col).value for col in range(1, 7)] == [
        'No.', 'Description & Specifications', 'Qty', 'Unit Price', 'Amount', 'Remarks'
    ]
    assert ws.freeze_panes == f'A{header_row + 1}'
    assert ws.print_area, "Excel 源文件必须设置打印区域，才能稳定导出 PDF"
    assert ws.page_setup.fitToWidth == 1


def test_sample_script_generates_matching_excel_source(tmp_path):
    """生成对外 HTML/PDF 时应同步生成同名 Excel 源文件"""
    fixture_path = os.path.join(FIXTURES_DIR, 'sample_customer.json')
    output_html = tmp_path / 'SPL-20260328-001.html'
    script_path = os.path.join(os.path.dirname(__file__), '..', '..', 'scripts', 'generate_sample.py')

    completed = subprocess.run(
        [sys.executable, script_path, '--data', fixture_path, '--output', str(output_html)],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr or completed.stdout
    assert output_html.exists()
    assert output_html.with_suffix('.xlsx').exists()


def test_sample_excel_handles_legacy_requirements_payload():
    """旧投递台生成的 requirements 样品数据也应走同一套对外源 Excel 版式"""
    data = {
        "meta": {
            "created_at": "2026-05-28",
            "target_company": "SNK-S (OOO СНК-С)",
        },
        "contact": {"name": "A. Ulyashkin", "email": "a.ulyashkin@snk-s.ru"},
        "requirements": [
            {
                "model": "AP-083A",
                "connector_type": "HDMI-HDMI",
                "wire_gauge": "HDMI2.1 30AWG CCS",
                "jacket_color": "Black",
                "requirements": "8K 60Hz, eARC",
                "length": "2",
                "quantity": 5,
                "notes": "Need certification",
            }
        ],
    }
    output_path = os.path.join(os.path.dirname(__file__), 'output', 'test-sample-legacy.xlsx')

    result = generate_sample_excel(data, output_path)
    assert result.success, f"生成应该成功：{result.error}"
    wb = load_workbook(output_path, data_only=False)
    ws = wb.active

    assert ws['E1'].value == 'SAMPLE REQUEST'
    assert find_cell_value(ws, 'SNK-S (OOO СНК-С)') is not None
    assert find_cell_containing(ws, 'AP-083A HDMI-HDMI') is not None


def test_sample_number_format():
    """样品单编号格式应该是 SPL-YYYYMMDD-XXX"""
    data = load_fixture('sample_customer.json')
    sample_no = data['sample']['sample_no']
    assert re.match(r'^SPL-\d{8}-\d{3}$', sample_no), f"样品单编号格式错误：{sample_no}"


def test_sample_quantity_reasonable():
    """样品数量应该合理（通常 <= 10）"""
    data = load_fixture('sample_customer.json')
    for product in data['products']:
        assert product['quantity'] <= 10, f"样品数量应该 <= 10: {product['quantity']}"


def test_freight_collect():
    """样品单应该指定运费到付"""
    data = load_fixture('sample_customer.json')
    assert data['shipping']['freight_collect'] == True
