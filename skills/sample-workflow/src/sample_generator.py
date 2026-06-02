#!/usr/bin/env python3
"""
样品单 Excel 源文件生成器。

这份 Excel 是对外 Sample Request PDF 的可编辑源文件，结构应与
scripts/generate_sample.py 生成的 HTML/PDF 保持一致。
"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass
class GenerationResult:
    success: bool
    output_path: Optional[str] = None
    error: Optional[str] = None


COLORS = {
    "dark": "111827",
    "light": "F3F4F6",
    "paper": "FFFFFF",
}


def border(style="thin"):
    side = Side(style=style, color="000000")
    return Border(top=side, left=side, right=side, bottom=side)


def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def money_format(currency: str):
    symbol = "$" if currency == "USD" else f'"{currency} "'
    return f'{symbol}#,##0.00'


def as_money(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def as_date(value):
    return str(value or datetime.now().strftime("%Y-%m-%d")).replace("-", ".")


def normalize_products(data: dict):
    products = data.get("products")
    if products:
        return products

    requirements = data.get("requirements", [])
    normalized = []
    for item in requirements:
        model = item.get("model", "")
        connector_type = item.get("connector_type", "")
        requirements_text = item.get("requirements", "")
        length = item.get("length", "")
        description = " ".join(part for part in [model, connector_type] if part).strip() or item.get("description", "")
        spec_parts = [
            item.get("wire_gauge", ""),
            item.get("connector_material", ""),
            f"Color: {item.get('jacket_color')}" if item.get("jacket_color") else "",
            f"Length: {length}m" if length else "",
            requirements_text,
            f"Printing: {item.get('printing_text')}" if item.get("printing_text") else "",
            f"Packaging: {item.get('packaging')}" if item.get("packaging") else "",
        ]
        normalized.append({
            "description": description,
            "specification": ", ".join(part for part in spec_parts if part),
            "quantity": item.get("quantity", 0),
            "unit_price": item.get("unit_price", 0),
            "remarks": item.get("notes", ""),
        })
    return normalized


def normalize_customer(data: dict):
    customer = data.get("customer")
    if customer:
        return customer

    meta = data.get("meta", {})
    return {
        "company_name": meta.get("target_company", ""),
        "contact": data.get("contact", {}).get("name", ""),
        "email": data.get("contact", {}).get("email", ""),
        "phone": data.get("contact", {}).get("phone", ""),
        "address": data.get("shipping_address", {}).get("address", ""),
        "country": data.get("shipping_address", {}).get("country", ""),
    }


def normalize_sample_info(data: dict):
    sample = data.get("sample")
    if sample:
        return sample

    meta = data.get("meta", {})
    created_at = meta.get("created_at", datetime.now().strftime("%Y-%m-%d"))
    compact_date = str(created_at).replace("-", "")
    return {
        "sample_no": data.get("sample_no", f"SPL-{compact_date}-001"),
        "date": created_at,
        "purpose": data.get("purpose", "Sample Request"),
    }


def write_label_value(ws, row, label, value, label_col=1, value_col=2, width=6):
    ws.cell(row=row, column=label_col, value=label)
    ws.cell(row=row, column=value_col, value=value)
    ws.cell(row=row, column=label_col).font = Font(bold=True)
    ws.cell(row=row, column=label_col).alignment = Alignment(vertical="top")
    ws.cell(row=row, column=value_col).alignment = Alignment(wrap_text=True, vertical="top")
    if width > 2:
        ws.merge_cells(
            start_row=row,
            start_column=value_col,
            end_row=row,
            end_column=label_col + width - 1,
        )
    for col in range(label_col, label_col + width):
        ws.cell(row=row, column=col).border = border()


def section_title(ws, row, title, last_col=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=11, color="000000")
    cell.fill = fill(COLORS["light"])
    cell.alignment = Alignment(vertical="center")
    cell.border = border()
    ws.row_dimensions[row].height = 22


def style_table_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = fill(COLORS["dark"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()


def style_table_cell(cell, horizontal="left", bold=False, number_format=None):
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=True)
    cell.border = border()
    if number_format:
        cell.number_format = number_format


def generate_sample_excel(data: dict, output_path: str) -> GenerationResult:
    """生成与对外 PDF 版式匹配的样品单 Excel 源文件"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sample Request"
        ws.sheet_view.showGridLines = False

        for column, width in {
            "A": 8,
            "B": 38,
            "C": 12,
            "D": 15,
            "E": 15,
            "F": 18,
        }.items():
            ws.column_dimensions[column].width = width

        company_name = data.get("company_name", "FARREACH ELECTRONIC CO LIMITED")
        company_address = data.get(
            "company_address",
            "No. 6, Chuangye Road East, Shuanglinpian, Liangang Industrial Zone, Zhuhai, China",
        )
        company_phone = data.get("company_phone", "YOUR-PHONE")
        company_fax = data.get("company_fax", "YOUR-FAX")
        company_website = data.get("company_website", "www.your-domain.com")
        company_email = data.get("company_email", "your-email")

        customer = normalize_customer(data)
        customer_name = customer.get("company_name", customer.get("name", "_________________"))
        customer_contact = customer.get("contact", customer.get("contact_name", "Procurement Manager"))
        customer_email = customer.get("email", "")
        customer_phone = customer.get("phone", "")

        sample = normalize_sample_info(data)
        sample_no = sample.get("sample_no", data.get("sampleNo", "SPL-" + datetime.now().strftime("%Y%m%d-001")))
        sample_date = as_date(sample.get("date", data.get("date")))
        purpose = sample.get("purpose", "Testing / Evaluation")

        shipping_addr = data.get("shipping_address", {})
        shipping_company = shipping_addr.get("company_name", customer_name)
        shipping_contact = shipping_addr.get("contact", customer_contact)
        shipping_phone = shipping_addr.get("phone", customer_phone)
        shipping_address = shipping_addr.get("address", customer.get("address", ""))
        shipping_country = shipping_addr.get("country", customer.get("country", ""))
        shipping_postal = shipping_addr.get("postal_code", "")

        products = normalize_products(data)
        currency = data.get("currency", "USD")
        freight = as_money(data.get("shipping", {}).get("freight_amount", 0))
        subtotal = sum(as_money(p.get("quantity")) * as_money(p.get("unit_price", p.get("unitPrice", 0))) for p in products)
        total = subtotal + freight

        shipping = data.get("shipping", {})
        shipping_method = shipping.get("method", "DHL")
        freight_account = shipping.get("account_no", "")
        freight_collect = shipping.get("freight_collect", True)

        terms = data.get("terms", {})
        sample_charge = terms.get("sample_charge", "Free")
        lead_time = terms.get("lead_time", "3-5 days after confirmation")
        remarks = terms.get("remarks", "")

        ws.merge_cells("A1:D3")
        ws["A1"] = company_name
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(vertical="top", wrap_text=True)

        ws.merge_cells("A4:D7")
        ws["A4"] = (
            f"Address: {company_address}\n"
            f"Tel: {company_phone} | Fax: {company_fax}\n"
            f"Email: {company_email}\n"
            f"Website: {company_website}"
        )
        ws["A4"].font = Font(size=10)
        ws["A4"].alignment = Alignment(vertical="top", wrap_text=True)

        ws.merge_cells("E1:F3")
        ws["E1"] = "SAMPLE REQUEST"
        ws["E1"].font = Font(bold=True, size=18)
        ws["E1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(1, 4):
            for col in range(5, 7):
                ws.cell(row=row, column=col).border = border("medium")

        for row in range(1, 8):
            ws.row_dimensions[row].height = 22

        row = 9
        write_label_value(ws, row, "Sample No:", sample_no, 1, 2, 3)
        write_label_value(ws, row, "Date:", sample_date, 4, 5, 3)
        row += 1
        write_label_value(ws, row, "Purpose:", purpose, 1, 2, 6)

        row += 2
        section_title(ws, row, "CUSTOMER INFORMATION:")
        row += 1
        write_label_value(ws, row, "Company:", customer_name)
        row += 1
        write_label_value(ws, row, "Attn:", customer_contact)
        row += 1
        write_label_value(ws, row, "Email:", customer_email)
        row += 1
        write_label_value(ws, row, "Phone:", customer_phone)

        row += 2
        section_title(ws, row, "SHIPPING ADDRESS:")
        row += 1
        write_label_value(ws, row, "Company:", shipping_company)
        row += 1
        write_label_value(ws, row, "Attn:", shipping_contact)
        row += 1
        write_label_value(ws, row, "Address:", f"{shipping_address}\n{shipping_country} {shipping_postal}".strip())
        ws.row_dimensions[row].height = 34
        row += 1
        write_label_value(ws, row, "Phone:", shipping_phone)

        row += 2
        section_title(ws, row, "SAMPLE ITEMS:")
        row += 1
        header_row = row
        headers = ["No.", "Description & Specifications", "Qty", "Unit Price", "Amount", "Remarks"]
        for col, header in enumerate(headers, start=1):
            style_table_header(ws.cell(row=header_row, column=col, value=header))
        ws.row_dimensions[header_row].height = 26

        first_product_row = header_row + 1
        row = first_product_row
        for idx, product in enumerate(products, start=1):
            description = product.get("description", "")
            specification = product.get("specification", "")
            quantity = as_money(product.get("quantity"))
            unit_price = as_money(product.get("unit_price", product.get("unitPrice", 0)))
            remarks_text = product.get("remarks", "")

            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=f"{description}\n{specification}".strip())
            ws.cell(row=row, column=3, value=quantity)
            ws.cell(row=row, column=4, value=unit_price)
            ws.cell(row=row, column=5, value=f"=C{row}*D{row}")
            ws.cell(row=row, column=6, value=remarks_text)

            style_table_cell(ws.cell(row=row, column=1), "center")
            style_table_cell(ws.cell(row=row, column=2))
            style_table_cell(ws.cell(row=row, column=3), "center")
            style_table_cell(ws.cell(row=row, column=4), "right", number_format=money_format(currency))
            style_table_cell(ws.cell(row=row, column=5), "right", bold=True, number_format=money_format(currency))
            style_table_cell(ws.cell(row=row, column=6))
            ws.row_dimensions[row].height = 46
            row += 1

        last_product_row = row - 1
        if not products:
            row += 1

        subtotal_row = row + 1
        ws.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=4)
        ws.cell(row=subtotal_row, column=1, value="Sample Subtotal:").alignment = Alignment(horizontal="right")
        ws.cell(row=subtotal_row, column=1).font = Font(bold=True)
        ws.cell(row=subtotal_row, column=5, value=f"=SUM(E{first_product_row}:E{row - 1})" if products else subtotal)

        freight_row = subtotal_row + 1
        ws.merge_cells(start_row=freight_row, start_column=1, end_row=freight_row, end_column=4)
        ws.cell(row=freight_row, column=1, value="Freight:").alignment = Alignment(horizontal="right")
        ws.cell(row=freight_row, column=1).font = Font(bold=True)
        ws.cell(row=freight_row, column=5, value=freight if freight > 0 else "Collect")

        total_row = freight_row + 1
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        ws.cell(row=total_row, column=1, value="TOTAL:").alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=total_row, column=5, value=f"=E{subtotal_row}+{freight}" if freight > 0 else total)
        ws.cell(row=total_row, column=5).font = Font(bold=True, size=12)

        for total_section_row in range(subtotal_row, total_row + 1):
            for col in range(1, 6):
                ws.cell(row=total_section_row, column=col).border = border()
                ws.cell(row=total_section_row, column=col).fill = fill(COLORS["light"]) if total_section_row == total_row else fill(COLORS["paper"])
            if isinstance(ws.cell(row=total_section_row, column=5).value, (int, float)) or str(ws.cell(row=total_section_row, column=5).value).startswith("="):
                ws.cell(row=total_section_row, column=5).number_format = money_format(currency)
            ws.cell(row=total_section_row, column=5).alignment = Alignment(horizontal="right")

        row = total_row + 2
        section_title(ws, row, "SHIPPING & TERMS:")
        row += 1
        write_label_value(ws, row, "Courier:", shipping_method)
        row += 1
        account_text = freight_account if freight_account else ("N/A (Freight Collect)" if freight_collect else "N/A")
        write_label_value(ws, row, "Freight Account:", account_text)
        row += 1
        write_label_value(ws, row, "Sample Charge:", sample_charge)
        row += 1
        write_label_value(ws, row, "Lead Time:", lead_time)

        if remarks:
            row += 2
            section_title(ws, row, "REMARKS:")
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1, value=remarks)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=1).border = border()
            ws.row_dimensions[row].height = 60

        last_row = row
        ws.freeze_panes = f"A{header_row + 1}"
        ws.print_area = f"A1:F{last_row}"
        ws.page_setup.paperSize = 9
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        ws.page_margins.left = 0.35
        ws.page_margins.right = 0.35
        ws.page_margins.top = 0.45
        ws.page_margins.bottom = 0.45
        ws.page_margins.header = 0
        ws.page_margins.footer = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.auto_filter.ref = f"A{header_row}:F{max(header_row, last_product_row)}"
        for row_idx in range(1, last_row + 1):
            ws.row_dimensions[row_idx].height = ws.row_dimensions[row_idx].height or 22

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        return GenerationResult(success=True, output_path=str(output))
    except Exception as e:
        return GenerationResult(success=False, error=str(e))
