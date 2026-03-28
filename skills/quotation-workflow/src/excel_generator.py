#!/usr/bin/env python3
"""
报价单 Excel 生成器
"""

from dataclasses import dataclass
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os


@dataclass
class GenerationResult:
    """生成结果"""
    success: bool
    output_path: Optional[str] = None
    error: Optional[str] = None


def generate_quotation_excel(data: dict, output_path: str) -> GenerationResult:
    """
    生成报价单 Excel 文件
    
    Args:
        data: 报价单数据（包含 customer, products, quotation, trade_terms）
        output_path: 输出文件路径
    
    Returns:
        GenerationResult: 生成结果
    """
    try:
        # 创建 Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Quotation'
        
        # 设置列宽
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # 样式
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center')
        right_align = Alignment(horizontal='right')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题行
        ws.merge_cells('A1:F1')
        title = ws['A1']
        title.value = 'QUOTATION'
        title.font = Font(bold=True, size=16)
        title.alignment = Alignment(horizontal='center')
        
        # 公司信息
        ws['A2'] = 'FARREACH ELECTRONIC CO LIMITED'
        ws['A2'].font = Font(bold=True)
        
        # 报价单信息
        row = 4
        ws[f'A{row}'] = f"Quotation No: {data.get('quotation', {}).get('quotation_no', 'N/A')}"
        ws[f'B{row}'] = f"Date: {data.get('quotation', {}).get('date', 'N/A')}"
        
        # 客户信息
        row = 6
        ws[f'A{row}'] = 'To:'
        ws[f'A{row}'].font = bold_font
        customer = data.get('customer', {})
        ws[f'A{row+1}'] = customer.get('company_name', '')
        ws[f'A{row+2}'] = customer.get('address', '')
        ws[f'A{row+3}'] = f"Tel: {customer.get('phone', '')}"
        ws[f'A{row+4}'] = f"Email: {customer.get('email', '')}"
        
        # 产品表格标题
        row = 12
        headers = ['No.', 'Description', 'Specification', 'Qty', 'Unit Price', 'Total']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # 产品数据
        products = data.get('products', [])
        for i, product in enumerate(products, start=1):
            row += 1
            ws.cell(row=row, column=1, value=i).alignment = center_align
            ws.cell(row=row, column=2, value=product.get('description', '')).border = thin_border
            ws.cell(row=row, column=3, value=product.get('specification', '')).border = thin_border
            ws.cell(row=row, column=4, value=product.get('quantity', 0)).alignment = center_align
            ws.cell(row=row, column=4).border = thin_border
            
            unit_price = product.get('unit_price', 0)
            ws.cell(row=row, column=5, value=f"{unit_price:.2f}").alignment = right_align
            ws.cell(row=row, column=5).border = thin_border
            
            line_total = product.get('quantity', 0) * unit_price
            ws.cell(row=row, column=6, value=f"{line_total:.2f}").alignment = right_align
            ws.cell(row=row, column=6).border = thin_border
        
        # 总额
        row += 1
        subtotal = sum(p.get('quantity', 0) * p.get('unit_price', 0) for p in products)
        ws.cell(row=row, column=5, value='TOTAL:').font = bold_font
        ws.cell(row=row, column=5).alignment = right_align
        ws.cell(row=row, column=6, value=f"{subtotal:.2f}").font = bold_font
        ws.cell(row=row, column=6).alignment = right_align
        
        # 确保输出目录存在
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        
        # 保存文件
        wb.save(output_path)
        
        return GenerationResult(
            success=True,
            output_path=output_path
        )
        
    except Exception as e:
        return GenerationResult(
            success=False,
            error=str(e)
        )
